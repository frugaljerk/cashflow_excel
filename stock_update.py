#! python3
# stock_update.py - read CSV files from TD Webroker and update stock holding's market value to current month.
import openpyxl, csv, os, datetime
from openpyxl.styles import Font, Alignment
import re

def main():
    #COPY a sheet from template/previous month sheet in Cashflow.xlsx
    wb = openpyxl.load_workbook('CashFlow 2021.xlsx')
    print(wb.sheetnames)
    lastsheet = wb.sheetnames[len(wb.sheetnames) - 1]
    sheet = wb.copy_worksheet(wb[lastsheet])
    bold = Font(bold=True)
    right_align = Alignment(horizontal='right')

    US_EXCHANGE = sheet['I3'].value
    #Rename active sheet to current month
    sheet.title = datetime.datetime.now().strftime("%B")

    #Keywords for dividend description in activity sheet
    ACTIVITY_KEYWORDS = ['DIV', 'TXPDDV', 'CIL']


    #Read all the CSV files one at a time from folder
    ROW_NUM = 32
    COL_NUM = 1

    dividend_sum = 0
    os.chdir(r'.\stockCSV')
    print(os.getcwd())
    for filename in os.listdir('.'):

        tmplist = [] #temp list for holding data for each filename
        #opening filename with 'holdings' to update stock market value
        if re.search('holdings', filename) != None:
            with open(filename, newline='') as csvfile:
                print(f'opening {filename}....')
                csvreader = csv.reader(csvfile)
                for row in csvreader:
                   tmplist.append(row)

            # Base on the accont type, insert Stock name (Col C) and Market value(Col H) to CashFlow 2021.xlsx
            sheet.cell(row=ROW_NUM, column=COL_NUM).value = tmplist[1][1] #assigning account name
            sheet.cell(row=ROW_NUM, column=COL_NUM).font = bold
            ROW_NUM += 1
            if tmplist[8][1] == 'CA':
                sheet.cell(row=ROW_NUM, column=COL_NUM).value = tmplist[2][0] #assigning cash balance
                sheet.cell(row=ROW_NUM, column=COL_NUM + 1).value = float(tmplist[2][1])
                ROW_NUM += 1
                for i in range(8, len(tmplist)):
                    sheet.cell(row=ROW_NUM, column=COL_NUM).value = tmplist[i][2]
                    sheet.cell(row=ROW_NUM, column=COL_NUM + 1).value = float(tmplist[i][7])
                    ROW_NUM += 1 #incrementing one row in xlsx after each value update
                #add two rows of padding before updating xlsx for next account
                ROW_NUM = ROW_NUM + 2
            # check if US dollar account
            elif tmplist[8][1] == 'US': #assign amount to US dollar column
                sheet.cell(row=ROW_NUM, column=COL_NUM).value = tmplist[2][0] #assigning cash balance
                sheet.cell(row=ROW_NUM, column=COL_NUM + 2).value = tmplist[2][1] #US currency
                sheet.cell(row=ROW_NUM, column=COL_NUM + 1).value = float(tmplist[2][1]) / US_EXCHANGE #CAD Currency
                ROW_NUM += 1
                for i in range(8, len(tmplist)):
                    sheet.cell(row=ROW_NUM, column=COL_NUM).value = tmplist[i][2]
                    sheet.cell(row=ROW_NUM, column=COL_NUM + 2).value = tmplist[i][7] #US Currency
                    sheet.cell(row=ROW_NUM, column=COL_NUM + 1).value = float(tmplist[i][7]) / US_EXCHANGE #CAD Currency
                    ROW_NUM += 1 #incrementing one row in xlsx after each value update
                #add two rows of padding before updating xlsx for next account
                ROW_NUM = ROW_NUM + 2

        #add the dividend up for the month from activity files
        elif re.search('activity', filename) != None:
            with open(filename, newline='') as csvfile:
                print(f'opening {filename}....')
                csvreader = csv.reader(csvfile)
                for row in csvreader:
                    tmplist.append(row)
            for i in range(4, len(tmplist)):
                #if month string(ie Jul) in trade date month and trade amount > 0, append to dividend
                if re.search(datetime.datetime.now().strftime("%b"),tmplist[i][0]) != None and tmplist[i][3] in ACTIVITY_KEYWORDS:
                    dividend_sum += float(tmplist[i][7])
                    print(f'{tmplist[i][2]} dividend added: {tmplist[i][7]}')



    #update the monthly dividend
    print(f'dividend sum: {dividend_sum}')
    sheet.cell(row=6, column=2).value = float(dividend_sum)
    os.chdir('..')
    #save .xlsx

    wb.save('CashFlow 2021.xlsx')

#TODO: USE API TO UPDATE CURRENCY and GOLD Price
#TODO: Utilitize Function oriented program